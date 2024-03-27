from flask import Blueprint, render_template, request

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.common.exceptions import NoSuchElementException, NoSuchWindowException
import openpyxl

from src.utils import formHandler
from src.utils import worksheetWritter

validator_page = Blueprint('validator_page',__name__, template_folder='templates')

@validator_page.route('/')
def main_page():
   return render_template('index.html')

@validator_page.route('/validar-datos', methods=['POST'])
def validar_uno():

   try:
      web = "https://dgii.gov.do/vehiculosMotor/consultas/Paginas/consultaPlacas.aspx"
      browser = webdriver.Chrome(
         service=Service(ChromeDriverManager().install()))

      placa = request.form["placa"]
      cedula = request.form["cedula"]
      vehiculos_list  = [formHandler.fill_in_form(browser,web,placa=placa,cedula=cedula)]
      
      browser.quit()
      
      excel_dataframe = openpyxl.load_workbook("src/db/vehiculosAudtoria.xlsx")

      data_frame = excel_dataframe.active

      exist_in_sheet = worksheetWritter.sheet_writter(vehiculos_list,data_frame=data_frame,cedula=cedula)
      
      excel_dataframe.save("src/db/vehiculosAudtoria.xlsx")
      excel_dataframe.close()
      return render_template('table.html', vehiculos_list=vehiculos_list, done=exist_in_sheet)
   except NoSuchElementException:
      return render_template("error.html", razon="Cedula o Placa Inválida.", solucion="Verificar los datos insertados en el formulario.")
   except PermissionError:
      return render_template("error.html",razon="No tiene permisos de escritura de la base de datos (excel)",solucion="En caso de tener el excel abierto, debe de cerrrarlo dado que no es posible actualizar los registros mientras esté abierto el archivo.")

   except NoSuchWindowException:
      return render_template("error.html",razon="El navegador se ha cerrado de forma inesperada.",solucion="Se debe de esperar a que los procesos se completen y no cerrar el navegador.")

   except Exception as ex:
      return render_template("error.html",razon=f"{type(ex).__name__}: {ex}",solucion="Reintentar y en caso del error persistir, realizar un reporte del mismo.")
    

@validator_page.route('/validar-datos/todos', methods=['GET'])
def validar_todos():

   try:
      web = "https://dgii.gov.do/vehiculosMotor/consultas/Paginas/consultaPlacas.aspx"
      browser = webdriver.Chrome(
        service=Service(ChromeDriverManager().install()))
      
      excel_dataframe = openpyxl.load_workbook("src/db/vehiculosAudtoria.xlsx")

      dataFrame = excel_dataframe.active
      
      vehiculos_list = []

      for row in range(2,dataFrame.max_row + 1):
         
         result = formHandler.fill_in_form(browser,web,dataFrame=dataFrame,row=row)
         
         vehiculos_list.append(result)
         
         
         worksheetWritter.sheet_writter(vehiculos_list,data_frame=dataFrame,row=row)
         
      browser.quit()

      excel_dataframe.save("src/db/vehiculosAudtoria.xlsx")
      excel_dataframe.close()
      return render_template('table.html', vehiculos_list=vehiculos_list,done=True)
   except NoSuchElementException:
      return render_template("error.html", razon="Cedula o Placa Inválida.", solucion="Verificar los datos insertados en el formulario.")
   except PermissionError:
      return render_template("error.html",razon="No tiene permisos de escritura de la base de datos (excel)",solucion="En caso de tener el excel abierto, debe de cerrrarlo dado que no es posible actualizar los registros mientras esté abierto el archivo.")

   except NoSuchWindowException:
      return render_template("error.html",razon="El navegador se ha cerrado de forma inesperada.",solucion="Se debe de esperar a que los procesos se completen y no cerrar el navegador.")

   except Exception as ex:
      print(type(ex).__name__)
      return render_template("error.html",razon=f"{type(ex).__name__}: {ex}",solucion="Reintentar y en caso del error persistir, realizar un reporte del mismo.")
    

@validator_page.route('/reporte')
def reporte_datos():
    
   vehiculos_list = []
   excel_dataframe = openpyxl.load_workbook("src/db/vehiculosAudtoria.xlsx")
    
   dataFrame = excel_dataframe.active
    
   for row in range(1,dataFrame.max_row):
      _row = [row,]
      for col in dataFrame.iter_cols(1,dataFrame.max_column):
         _row.append(col[row].value)
      vehiculos_list.append(_row)
    
   excel_dataframe.close()
   return render_template('reporte.html', vehiculos_list=vehiculos_list)
